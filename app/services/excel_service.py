import os
import re
import tempfile
import time
import uuid
from collections.abc import Callable
from datetime import date, datetime

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.config import env
from app.providers.s3_provider import build_object_url
from app.services.reconciliation_service import _split_object_reference
from app.types import DoCleanupResult, ReconciliationResult, StorageSummary

_BOLD = Font(bold=True)
_RED = Font(color="FFCC0000")
# Applied directly instead of `cell.style = "Hyperlink"` — the named-style path re-resolves the
# style from the workbook's style registry on every single cell, which measurably adds up across
# tens/hundreds of thousands of rows. A plain Font gets the same visual result (and still a real,
# clickable hyperlink) for a fraction of the per-cell cost.
_HYPERLINK_FONT = Font(color="0563C1", underline="single")

# How often generate_reconciliation_report reports progress while writing rows — frequent enough
# to feel alive on a huge report, not so frequent that the progress calls themselves add overhead.
_REPORT_TICK_EVERY_ROWS = 2000

ReportProgressCallback = Callable[[str, int, int, str, str | None], None]


def _format_bytes(num_bytes: float) -> str:
    if num_bytes == 0:
        return "0 B"
    units = ["B", "KB", "MB", "GB", "TB"]
    exponent = 0
    value = float(num_bytes)
    while value >= 1024 and exponent < len(units) - 1:
        value /= 1024
        exponent += 1
    return f"{value:.2f} {units[exponent]}"


def _mb(num_bytes: float) -> float:
    return round(num_bytes / (1024 * 1024), 2)


def _gb(num_bytes: float) -> float:
    return round(num_bytes / (1024 * 1024 * 1024), 2)


def _naive(dt: datetime | None) -> datetime | None:
    """openpyxl/Excel cells can't hold a timezone-aware datetime ('Excel does not support
    timezones in datetimes') — boto3 always returns LastModified as UTC-aware, so this strips
    tzinfo while keeping the same UTC wall-clock value."""
    return dt.replace(tzinfo=None) if dt is not None else None


def _year_month(dt: datetime | None) -> str:
    return dt.strftime("%Y-%m") if dt is not None else ""


def _style_header_row(sheet: Worksheet) -> None:
    for cell in sheet[1]:
        cell.font = _BOLD


def _today_date_stamp() -> str:
    return date.today().isoformat()


def buckets_label(buckets: list[str] | None) -> str:
    """One bucket -> its name; several -> joined; none/"all buckets" -> "all-buckets"."""
    if not buckets:
        return "all-buckets"
    if len(buckets) <= 3:
        return "+".join(buckets)
    return f"{len(buckets)}-buckets"


def build_report_file_name(parts: list[str | None]) -> str:
    """Builds a `part-part-...-YYYY-MM-DD.xlsx` filename, skipping empty parts."""
    segments = []
    for part in parts:
        if not part:
            continue
        cleaned = re.sub(r"[^a-zA-Z0-9_-]+", "-", part).strip("-")
        if cleaned:
            segments.append(cleaned)
    segments.append(_today_date_stamp())
    return f"{'-'.join(segments)}.xlsx"


def _save_workbook(workbook: Workbook, file_name: str) -> str:
    """
    Writes to a temp file and renames it into place, so a concurrent read (e.g. downloading a
    partial report mid-scan) always sees either the previous or the fully-written new version,
    never a torn/corrupt file from an in-progress write.
    """
    os.makedirs(env.reports_dir, exist_ok=True)
    file_path = os.path.join(env.reports_dir, file_name)
    tmp_path = os.path.join(env.reports_dir, f".{file_name}.{uuid.uuid4().hex}.tmp")
    workbook.save(tmp_path)
    os.replace(tmp_path, file_path)
    return file_name


def generate_storage_report(summary: StorageSummary, file_name: str | None = None) -> str:
    file_name = file_name or build_report_file_name(["storage"])
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Storage Summary"

    headers = ["Bucket", "Object Count", "Total Size (Bytes)", "Total Size", "Error"]
    sheet.append(headers)
    widths = [30, 16, 20, 16, 40]
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=idx).column_letter].width = width

    for bucket in summary.buckets:
        sheet.append(
            [
                bucket.bucket,
                bucket.object_count,
                bucket.total_size,
                _format_bytes(bucket.total_size),
                bucket.error or "",
            ]
        )
        if bucket.error:
            sheet.cell(row=sheet.max_row, column=5).font = _RED

    sheet.append([])
    sheet.append(
        ["TOTAL", summary.object_count, summary.total_size, _format_bytes(summary.total_size), ""]
    )
    for cell in sheet[sheet.max_row]:
        cell.font = _BOLD

    _style_header_row(sheet)

    return _save_workbook(workbook, file_name)


def generate_copy_report(entries: list[dict], dest_provider: str | None, file_name: str | None = None) -> str:
    """
    One row per item a /storage/copy(/stream) run was asked to handle, deliverable for "what
    actually got moved to the new provider and where does it live now" — each entry is:
    {sourcePath, bucket, key, destBucket, sizeMb, table, column, rowId, status, error} where status is "Copied"
    (freshly transferred), "Skipped" (already existed at the destination — see copy_objects'
    overwrite=False default), or "Failed". The Destination URL column is built the same way
    every other report's clickable link is (build_object_url) and is shown even for a Failed row
    (where it points at where the object would be) so a reviewer can immediately tell the two
    providers' copies apart without reconstructing the URL by hand.
    """
    file_name = file_name or build_report_file_name(["copy"])
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Copied"

    headers = [
        "Source Path", "Bucket", "Key", "Destination URL", "Size (MB)",
        "Table", "Column", "Row ID", "Status", "Error",
    ]
    sheet.append(headers)
    widths = [55, 20, 45, 55, 12, 24, 24, 16, 12, 30]
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=idx).column_letter].width = width

    status_counts: dict[str, int] = {}
    for entry in entries:
        status = entry["status"]
        status_counts[status] = status_counts.get(status, 0) + 1

        dest_url = build_object_url(dest_provider, entry["destBucket"], entry["key"]) if entry.get("destBucket") else None
        row_idx = sheet.max_row + 1
        sheet.append(
            [
                entry.get("sourcePath") or "",
                entry["bucket"],
                entry["key"],
                dest_url or "",
                entry.get("sizeMb"),
                entry.get("table") or "",
                entry.get("column") or "",
                entry.get("rowId") if entry.get("rowId") is not None else "",
                status,
                entry.get("error") or "",
            ]
        )
        if entry.get("sourcePath"):
            cell = sheet.cell(row=row_idx, column=1)
            cell.hyperlink = entry["sourcePath"]
            cell.font = _HYPERLINK_FONT
        if dest_url:
            cell = sheet.cell(row=row_idx, column=4)
            cell.hyperlink = dest_url
            cell.font = _HYPERLINK_FONT
        if status == "Failed":
            sheet.cell(row=row_idx, column=9).font = _RED

    _style_header_row(sheet)

    summary_sheet = workbook.create_sheet("Summary")
    summary_sheet.append(["Status", "Count"])
    for status, count in status_counts.items():
        summary_sheet.append([status, count])
    _style_header_row(summary_sheet)
    for idx, width in enumerate([16, 10], start=1):
        summary_sheet.column_dimensions[get_column_letter(idx)].width = width

    return _save_workbook(workbook, file_name)


def parse_copy_report_for_db_update(file_obj) -> tuple[list[dict], dict[str, int]]:
    """
    Reads a generated migration report and returns only successfully available destination URLs.
    Failed rows are ignored; Copied and Skipped are both eligible because Skipped means the object
    was already present at the destination. Table/Column/Row ID are mandatory for a safe update.
    """
    try:
        workbook = load_workbook(file_obj, read_only=True, data_only=True)
    except Exception as error:  # noqa: BLE001 - openpyxl uses several zip/XML exception types
        raise ValueError(f"Not a valid .xlsx file: {error}") from error

    if "Copied" not in workbook.sheetnames:
        raise ValueError("No 'Copied' sheet found — upload a report generated by storage migration")

    rows_iter = workbook["Copied"].iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        return [], {"total": 0, "eligible": 0, "failed": 0, "incomplete": 0}
    col_index = {str(name).strip().lower(): i for i, name in enumerate(header) if name is not None}
    required = ["source path", "destination url", "table", "column", "row id", "status"]
    missing = [name for name in required if name not in col_index]
    if missing:
        raise ValueError(f"'Copied' sheet is missing expected column(s): {', '.join(missing)}")

    def cell_value(row, column_name: str):
        index = col_index[column_name]
        return row[index] if index < len(row) else None

    results = []
    stats = {"total": 0, "eligible": 0, "failed": 0, "incomplete": 0}
    for row in rows_iter:
        if not any(value is not None for value in row):
            continue
        stats["total"] += 1
        status = str(cell_value(row, "status") or "").strip().lower()
        if status not in {"copied", "skipped"}:
            stats["failed"] += 1
            continue
        item = {
            "sourcePath": cell_value(row, "source path"),
            "destinationUrl": cell_value(row, "destination url"),
            "table": cell_value(row, "table"),
            "column": cell_value(row, "column"),
            "rowId": cell_value(row, "row id"),
        }
        if any(value is None or value == "" for value in item.values()):
            stats["incomplete"] += 1
            continue
        results.append(item)
        stats["eligible"] += 1
    return results, stats


def parse_copy_report_for_delete(file_obj) -> tuple[list[dict], dict[str, int]]:
    """Extracts source objects that successfully reached the destination from a copy report."""
    results, stats, _excluded = parse_copy_report_for_delete_details(file_obj)
    return results, stats


def parse_copy_report_for_delete_details(file_obj) -> tuple[list[dict], dict[str, int], list[dict]]:
    """Extracts deletable source objects plus every rejected row and its reason."""
    try:
        workbook = load_workbook(file_obj, read_only=True, data_only=True)
    except Exception as error:  # noqa: BLE001 - openpyxl uses several zip/XML exception types
        raise ValueError(f"Not a valid .xlsx file: {error}") from error

    if "Copied" not in workbook.sheetnames:
        raise ValueError("No 'Copied' sheet found — upload a report generated by storage migration")

    rows_iter = workbook["Copied"].iter_rows(values_only=True)
    header = next(rows_iter, None)
    empty_stats = {"total": 0, "eligible": 0, "copied": 0, "skipped": 0, "failed": 0, "incomplete": 0}
    if not header:
        return [], empty_stats, []

    col_index = {str(name).strip().lower(): i for i, name in enumerate(header) if name is not None}
    required = ["bucket", "key", "destination url", "status"]
    missing = [name for name in required if name not in col_index]
    if missing:
        raise ValueError(f"'Copied' sheet is missing expected column(s): {', '.join(missing)}")

    def cell_value(row, column_name: str):
        index = col_index.get(column_name)
        return row[index] if index is not None and index < len(row) else None

    results = []
    excluded = []
    stats = empty_stats.copy()
    for row in rows_iter:
        if not any(value is not None for value in row):
            continue
        stats["total"] += 1
        status = str(cell_value(row, "status") or "").strip().lower()
        if status not in {"copied", "skipped"}:
            stats["failed"] += 1
            excluded.append(
                {
                    "bucket": cell_value(row, "bucket"),
                    "key": cell_value(row, "key"),
                    "destinationUrl": cell_value(row, "destination url"),
                    "status": str(cell_value(row, "status") or ""),
                    "reason": cell_value(row, "error") or "Migration status is not Copied or Skipped",
                }
            )
            continue

        bucket = cell_value(row, "bucket")
        key = cell_value(row, "key")
        destination_url = cell_value(row, "destination url")
        if not bucket or not key or not destination_url:
            stats["incomplete"] += 1
            missing = [
                label
                for label, value in (("Bucket", bucket), ("Key", key), ("Destination URL", destination_url))
                if not value
            ]
            excluded.append(
                {
                    "bucket": bucket,
                    "key": key,
                    "destinationUrl": destination_url,
                    "status": status.capitalize(),
                    "reason": f"Missing required value(s): {', '.join(missing)}",
                }
            )
            continue

        results.append(
            {
                "bucket": bucket,
                "key": key,
                "destinationUrl": destination_url,
                "sizeMb": cell_value(row, "size (mb)"),
                "status": status.capitalize(),
            }
        )
        stats[status] += 1
        stats["eligible"] += 1
    return results, stats, excluded


def generate_deletion_report(
    requested: list[dict],
    results: list[dict],
    excluded: list[dict] | None = None,
    file_name: str | None = None,
) -> str:
    """Writes final deletion outcomes and pre-delete exclusions to one workbook."""
    file_name = file_name or f"deletion-{datetime.now():%Y%m%dT%H%M%S}-{uuid.uuid4().hex[:8]}.xlsx"
    workbook = Workbook(write_only=True)

    headers = ["Bucket", "Key", "Path", "Size (MB)", "Error"]
    deleted_sheet = workbook.create_sheet("Deleted")
    failed_sheet = workbook.create_sheet("Failed")
    deleted_sheet.append(_header_row(deleted_sheet, headers))
    failed_sheet.append(_header_row(failed_sheet, headers))

    deleted_count = 0
    failed_count = 0
    for index, result in enumerate(results):
        request = requested[index] if index < len(requested) else {}
        row = [
            result.get("bucket") or request.get("bucket") or "",
            result.get("key") or request.get("key") or "",
            request.get("path") or "",
            request.get("sizeMb"),
            result.get("error") or "",
        ]
        if result.get("success"):
            deleted_sheet.append(row)
            deleted_count += 1
        else:
            failed_sheet.append(row)
            failed_count += 1

    excluded_rows = excluded or []
    excluded_sheet = workbook.create_sheet("Excluded")
    excluded_sheet.append(
        _header_row(excluded_sheet, ["Bucket", "Key", "Destination URL", "Status", "Reason"])
    )
    for item in excluded_rows:
        excluded_sheet.append(
            [
                item.get("bucket") or "",
                item.get("key") or "",
                item.get("destinationUrl") or "",
                item.get("status") or "",
                item.get("reason") or item.get("error") or "",
            ]
        )

    summary = workbook.create_sheet("Summary", 0)
    summary.append(_header_row(summary, ["Status", "Count"]))
    summary.append(["Deleted", deleted_count])
    summary.append(["Failed", failed_count])
    summary.append(["Excluded", len(excluded_rows)])
    return _save_workbook(workbook, file_name)


def _header_row(ws, headers: list[str]) -> list[WriteOnlyCell]:
    cells = []
    for h in headers:
        cell = WriteOnlyCell(ws, value=h)
        cell.font = _BOLD
        cells.append(cell)
    return cells


def _hyperlink_cell(ws, url: str) -> WriteOnlyCell:
    cell = WriteOnlyCell(ws, value=url)
    cell.hyperlink = url
    cell.font = _HYPERLINK_FONT
    return cell


def _raw_value_cell(ws, raw_value: str | None, fallback_url: str | None) -> WriteOnlyCell:
    """
    Displays exactly what's stored in the DB column (`raw_value`) — never a URL rebuilt from the
    parsed bucket/key, which can look nothing like the original (different host, different
    path-vs-virtual-hosted style, a CDN domain the app doesn't even know about) even though it
    points at the same object. `fallback_url` only supplies the *clickable* href: the raw value
    itself when it's already a full URL, otherwise a browsable link built from the resolved
    bucket/key so the cell stays clickable — the displayed text is untouched either way.
    """
    display = raw_value if raw_value is not None else (fallback_url or "")
    href = raw_value if raw_value and raw_value.lower().startswith(("http://", "https://")) else fallback_url
    cell = WriteOnlyCell(ws, value=display)
    if href:
        cell.hyperlink = href
        cell.font = _HYPERLINK_FONT
    return cell


def generate_reconciliation_report(
    result: ReconciliationResult,
    file_name: str | None = None,
    provider: str | None = None,
    on_progress: ReportProgressCallback | None = None,
) -> str:
    """
    Uses openpyxl's write_only mode: a normal Workbook re-registers a cell's style against the
    workbook's shared style table on every assignment, which gets measurably slower as that
    table grows — noticeable well before "thousands and thousands of objects" territory, since
    every Matched/Orphan row sets a font for its hyperlink. write_only sidesteps that (several
    times faster in practice), at the cost of being append-only — no going back to re-read or
    restyle a row after it's written, which this function never needs to do anyway.
    """
    file_name = file_name or build_report_file_name(["reconciliation"])
    workbook = Workbook(write_only=True)

    total_rows = len(result.matched) + len(result.missing) + len(result.orphan) + len(result.protected)
    written = 0
    started_at = time.monotonic()

    def tick(sheet_label: str) -> None:
        nonlocal written
        written += 1
        if on_progress and written % _REPORT_TICK_EVERY_ROWS == 0:
            elapsed = max(time.monotonic() - started_at, 0.001)
            on_progress(
                "report",
                written,
                total_rows,
                f"Writing {sheet_label} sheet — {written:,}/{total_rows:,} row(s) "
                f"({written / elapsed:,.0f}/s)",
                None,
            )

    matched_sheet = workbook.create_sheet("Matched")
    for idx, width in enumerate([60, 20, 14, 22, 16, 16, 12], start=1):
        matched_sheet.column_dimensions[get_column_letter(idx)].width = width
    matched_sheet.append(
        _header_row(
            matched_sheet, ["Value (as stored in DB)", "Bucket", "Size (MB)", "Last Modified", "Table", "Column", "Row ID"]
        )
    )
    for file in result.matched:
        fallback_url = build_object_url(provider, file.bucket, file.path)
        matched_sheet.append(
            [
                _raw_value_cell(matched_sheet, file.raw_value, fallback_url),
                file.bucket,
                _mb(file.size),
                _naive(file.last_modified),
                file.table,
                file.column,
                file.id,
            ]
        )
        tick("Matched")

    missing_sheet = workbook.create_sheet("Missing")
    for idx, width in enumerate([60, 16, 16, 12], start=1):
        missing_sheet.column_dimensions[get_column_letter(idx)].width = width
    missing_sheet.append(_header_row(missing_sheet, ["Value (as stored in DB)", "Table", "Column", "Row ID"]))
    for file in result.missing:
        # A missing file was never found in storage, so this link (when we have enough info to
        # build one at all) points at where it *would* be — it will 404, but that's still more
        # useful for tracking it down than a bare relative key. The displayed text is always the
        # raw DB value regardless — never a URL rebuilt from the parsed bucket/key.
        fallback_url = build_object_url(provider, file.bucket, file.path) if file.bucket else None
        missing_sheet.append(
            [_raw_value_cell(missing_sheet, file.raw_value, fallback_url), file.table, file.column, file.id]
        )
        tick("Missing")

    orphan_sheet = workbook.create_sheet("Orphan")
    for idx, width in enumerate([60, 14, 14], start=1):
        orphan_sheet.column_dimensions[get_column_letter(idx)].width = width
    # Bucket/Key are written but hidden — keeps the visible sheet exactly the 3 columns asked
    # for, while still letting the delete feature (parse_deletable_report) recover the exact
    # (bucket, key) pair for each row without depending on the clickable link's URL format.
    orphan_sheet.column_dimensions[get_column_letter(4)].hidden = True
    orphan_sheet.column_dimensions[get_column_letter(5)].hidden = True
    orphan_sheet.append(_header_row(orphan_sheet, ["Path", "Size (MB)", "Last Modified", "Bucket", "Key"]))
    for file in result.orphan:
        url = build_object_url(provider, file.bucket, file.path)
        orphan_sheet.append(
            [
                _hyperlink_cell(orphan_sheet, url),
                _mb(file.size),
                _year_month(file.last_modified),
                file.bucket,
                file.path,
            ]
        )
        tick("Orphan")

    # Only present when the run was given content_mappings (rich-text/JSON columns) to cross-check
    # against — otherwise Orphan already is the full "not found anywhere" set and this stays empty.
    # A file lands here instead of Orphan because its key was found embedded inside one of those
    # columns (e.g. a CKEditor/CKFinder upload referenced only as an <img> inside a lesson's HTML
    # body) — kept visible rather than silently dropped, so a reviewer can see why it was spared.
    protected_sheet = workbook.create_sheet("Protected (rich-text, JSON)")
    for idx, width in enumerate([60, 20, 14, 22], start=1):
        protected_sheet.column_dimensions[get_column_letter(idx)].width = width
    protected_sheet.append(_header_row(protected_sheet, ["Path", "Bucket", "Size (MB)", "Last Modified"]))
    for file in result.protected:
        url = build_object_url(provider, file.bucket, file.path)
        protected_sheet.append(
            [_hyperlink_cell(protected_sheet, url), file.bucket, _mb(file.size), _naive(file.last_modified)]
        )
        tick("Protected")

    if on_progress:
        on_progress("report", total_rows, total_rows, "Saving workbook to disk…", None)

    # Missing files were never found in storage, so they have no size to report; matched + orphan
    # + protected together account for every object actually seen in storage, so their sizes sum
    # to the same total as "Object Storage Objects".
    matched_size = sum(f.size for f in result.matched)
    orphan_size = sum(f.size for f in result.orphan)
    protected_size = sum(f.size for f in result.protected)

    summary_sheet = workbook.create_sheet("Summary")
    for idx, width in enumerate([34, 16, 16], start=1):
        summary_sheet.column_dimensions[get_column_letter(idx)].width = width
    summary_sheet.append(_header_row(summary_sheet, ["Metric", "Count", "Size (GB)"]))
    summary_sheet.append(["Matched", result.summary.matched_count, _gb(matched_size)])
    summary_sheet.append(["Missing", result.summary.missing_count, ""])
    summary_sheet.append(["Orphan (confirmed unused)", result.summary.orphan_count, _gb(orphan_size)])
    summary_sheet.append(["Protected (in rich-text/JSON content)", result.summary.protected_count, _gb(protected_size)])
    summary_sheet.append(["Database File References", result.summary.database_file_count, ""])
    summary_sheet.append(["Rich-text/JSON Rows Scanned", result.summary.content_reference_count, ""])
    summary_sheet.append(
        [
            "Object Storage Objects",
            result.summary.storage_object_count,
            _gb(matched_size + orphan_size + protected_size),
        ]
    )
    summary_sheet.append(["Skipped (other provider)", result.summary.other_provider_count, ""])
    summary_sheet.append(["Skipped (different provider host)", result.summary.different_provider_count, ""])

    return _save_workbook(workbook, file_name)


def generate_do_cleanup_report(
    result: DoCleanupResult,
    file_name: str | None = None,
    provider: str | None = None,
    on_progress: ReportProgressCallback | None = None,
) -> str:
    """
    Same write_only shape as generate_reconciliation_report. "Cleanup Candidates" is the actual
    deliverable (orphaned, and not found embedded in any rich-text/CKEditor column either);
    "Protected" is kept alongside for transparency — it's the reason a file that showed up as
    Orphan in a plain reconciliation report is *not* here, so a reviewer isn't left guessing.
    """
    file_name = file_name or build_report_file_name(["do-cleanup"])
    workbook = Workbook(write_only=True)

    total_rows = len(result.candidates) + len(result.protected)
    written = 0
    started_at = time.monotonic()

    def tick(sheet_label: str) -> None:
        nonlocal written
        written += 1
        if on_progress and written % _REPORT_TICK_EVERY_ROWS == 0:
            elapsed = max(time.monotonic() - started_at, 0.001)
            on_progress(
                "report",
                written,
                total_rows,
                f"Writing {sheet_label} sheet — {written:,}/{total_rows:,} row(s) ({written / elapsed:,.0f}/s)",
                None,
            )

    candidates_sheet = workbook.create_sheet("Cleanup Candidates")
    for idx, width in enumerate([60, 20, 14, 22], start=1):
        candidates_sheet.column_dimensions[get_column_letter(idx)].width = width
    # Key is written but hidden — Bucket is already a visible column here, but the delete
    # feature (parse_deletable_report) still needs the exact, unmodified key rather than the
    # clickable link's display URL.
    candidates_sheet.column_dimensions[get_column_letter(5)].hidden = True
    candidates_sheet.append(_header_row(candidates_sheet, ["Path", "Bucket", "Size (MB)", "Last Modified", "Key"]))
    for file in result.candidates:
        url = build_object_url(provider, file.bucket, file.path)
        candidates_sheet.append(
            [_hyperlink_cell(candidates_sheet, url), file.bucket, _mb(file.size), _naive(file.last_modified), file.path]
        )
        tick("Cleanup Candidates")

    protected_sheet = workbook.create_sheet("Protected (in rich-text content)")
    for idx, width in enumerate([60, 20, 14, 22], start=1):
        protected_sheet.column_dimensions[get_column_letter(idx)].width = width
    protected_sheet.append(_header_row(protected_sheet, ["Path", "Bucket", "Size (MB)", "Last Modified"]))
    for file in result.protected:
        url = build_object_url(provider, file.bucket, file.path)
        protected_sheet.append(
            [_hyperlink_cell(protected_sheet, url), file.bucket, _mb(file.size), _naive(file.last_modified)]
        )
        tick("Protected")

    if on_progress:
        on_progress("report", total_rows, total_rows, "Saving workbook to disk…", None)

    candidate_size = sum(f.size for f in result.candidates)
    protected_size = sum(f.size for f in result.protected)

    summary_sheet = workbook.create_sheet("Summary")
    for idx, width in enumerate([36, 16, 16], start=1):
        summary_sheet.column_dimensions[get_column_letter(idx)].width = width
    summary_sheet.append(_header_row(summary_sheet, ["Metric", "Count", "Size (GB)"]))
    summary_sheet.append(["Cleanup Candidates", result.summary.candidate_count, _gb(candidate_size)])
    summary_sheet.append(["Protected (in rich-text content)", result.summary.protected_count, _gb(protected_size)])
    summary_sheet.append(["Total Orphan (candidates + protected)", result.summary.orphan_count, ""])
    summary_sheet.append(["Matched", result.summary.matched_count, ""])
    summary_sheet.append(["Missing", result.summary.missing_count, ""])
    summary_sheet.append(["Database File References", result.summary.database_file_count, ""])
    summary_sheet.append(["Rich-text Rows Scanned", result.summary.content_reference_count, ""])
    summary_sheet.append(["Object Storage Objects", result.summary.storage_object_count, ""])
    summary_sheet.append(["Skipped (other provider)", result.summary.other_provider_count, ""])
    summary_sheet.append(["Skipped (different provider host)", result.summary.different_provider_count, ""])

    return _save_workbook(workbook, file_name)


_DELETABLE_SHEET_NAMES = ("Orphan", "Cleanup Candidates")


def parse_deletable_report(file_obj) -> list[dict]:
    """
    Reads a previously-downloaded Orphan/Cleanup Candidates report back in, for the delete
    feature: a human has already reviewed this exact file (possibly hand-edited — trimmed down
    to a shortlist, reordered, whatever) and is now selecting which rows to remove.

    Prefers the hidden Bucket/Key columns those sheets carry (see generate_reconciliation_report
    / generate_do_cleanup_report) since they're exact and don't care about column order. But a
    hand-edited copy easily loses hidden columns entirely (Excel rewrites the whole file on
    save), so when they're missing this falls back to re-deriving (bucket, key) from the visible
    Path column's URL — the same parsing reconciliation already relies on elsewhere, so it works
    for any row whose Path still looks like a URL this app generated. Raises ValueError on
    anything that doesn't look like one of those two reports, rather than silently returning
    nothing.
    """
    try:
        workbook = load_workbook(file_obj, read_only=True, data_only=True)
    except Exception as error:  # noqa: BLE001 - openpyxl raises its own zip/XML errors for anything not a real .xlsx
        raise ValueError(f"Not a valid .xlsx file: {error}") from error

    sheet_name = next((s for s in _DELETABLE_SHEET_NAMES if s in workbook.sheetnames), None)
    if sheet_name is None:
        raise ValueError(
            f"No {' or '.join(_DELETABLE_SHEET_NAMES)} sheet found in the uploaded file — "
            "upload a report downloaded from Auto reconciliation or DO cleanup candidates."
        )
    ws = workbook[sheet_name]

    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        return []
    col_index = {str(name).strip().lower(): i for i, name in enumerate(header) if name is not None}
    if "path" not in col_index or "size (mb)" not in col_index:
        raise ValueError(f"'{sheet_name}' sheet is missing expected column(s): path, size (mb)")
    has_bucket_key_columns = "bucket" in col_index and "key" in col_index

    results = []
    for row in rows_iter:
        path = row[col_index["path"]]
        if has_bucket_key_columns:
            bucket, key = row[col_index["bucket"]], row[col_index["key"]]
        elif path:
            bucket, key, _provider = _split_object_reference(str(path))
        else:
            bucket, key = None, None
        if not bucket or not key:
            continue
        last_modified = row[col_index["last modified"]] if "last modified" in col_index else None
        results.append(
            {
                "path": path,
                "bucket": bucket,
                "key": key,
                "sizeMb": row[col_index["size (mb)"]],
                "lastModified": str(last_modified) if last_modified is not None else None,
            }
        )
    return results


_MATCHED_SHEET_NAME = "Matched"
# The exported header has been "Path" since the write side moved to always showing the raw DB
# value under that name (see generate_reconciliation_report); older downloaded reports may still
# carry the previous header, so both are accepted here.
_MATCHED_PATH_HEADERS = ("path", "value (as stored in db)")


def parse_matched_report(file_obj) -> list[dict]:
    """
    Reads a previously-downloaded reconciliation report's "Matched" sheet back in — the source
    list for copying files to another provider (e.g. DO -> Wasabi) rather than deleting anything.
    Matched rows are files already confirmed to exist in both the DB and storage, so unlike
    parse_deletable_report there's no hidden Bucket/Key pair to prefer: the sheet's own visible
    Bucket column is trustworthy (it came straight from the storage listing during the scan), and
    the key is re-derived from the Path column's raw DB value the same way reconciliation itself
    resolved it (_split_object_reference) — a bare relative key, or a full URL through the
    provider or a CDN in front of it. Raises ValueError on anything that doesn't look like a
    reconciliation report.
    """
    try:
        workbook = load_workbook(file_obj, read_only=True, data_only=True)
    except Exception as error:  # noqa: BLE001 - openpyxl raises its own zip/XML errors for anything not a real .xlsx
        raise ValueError(f"Not a valid .xlsx file: {error}") from error

    if _MATCHED_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(
            f"No '{_MATCHED_SHEET_NAME}' sheet found in the uploaded file — upload a report "
            "downloaded from Auto reconciliation or Manual reconciliation."
        )
    ws = workbook[_MATCHED_SHEET_NAME]

    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        return []
    col_index = {str(name).strip().lower(): i for i, name in enumerate(header) if name is not None}

    path_col = next((h for h in _MATCHED_PATH_HEADERS if h in col_index), None)
    if path_col is None or "bucket" not in col_index:
        raise ValueError(f"'{_MATCHED_SHEET_NAME}' sheet is missing expected column(s): path, bucket")

    def cell_value(row, column_name: str):
        """Return None when an edited/blank Excel row ends before the requested column."""
        index = col_index.get(column_name)
        return row[index] if index is not None and index < len(row) else None

    results = []
    for row in rows_iter:
        raw_value = cell_value(row, path_col)
        bucket = cell_value(row, "bucket")
        if not raw_value or not bucket:
            continue
        _url_bucket, key, _provider_hint = _split_object_reference(str(raw_value))
        if not key:
            continue
        results.append(
            {
                "rawValue": raw_value,
                "bucket": bucket,
                "key": key,
                "sizeMb": cell_value(row, "size (mb)"),
                "lastModified": (
                    str(cell_value(row, "last modified")) if cell_value(row, "last modified") is not None else None
                ),
                "table": cell_value(row, "table"),
                "column": cell_value(row, "column"),
                "rowId": cell_value(row, "row id"),
            }
        )
    return results
